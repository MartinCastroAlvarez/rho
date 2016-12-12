from openpyxl.styles import Style, Font, Alignment
from lxml import etree
import cookielib
import requests
import openpyxl
import commands 
import getopt
import rho
import time
import sys

class MercadoLibre(rho.RhoHTTP):

	def search(self,keyword,limit=100):
		PAGE = 1
		urls = []
		results = []
		QUERY = keyword.upper().replace(" ","-")
		isNot404 = True
		while len(urls) < limit and isNot404:
			try: 
				url="http://listado.mercadolibre.com.ar/%s_Desde_%i" % (QUERY,PAGE)
				r = self.http("http://listado.mercadolibre.com.ar/%s_Desde_%i" % (QUERY,PAGE))
				root = etree.fromstring(r,parser=self.parser)
				for li in root.xpath('//*[@id="searchResults"]/li'):
					if len(urls) < limit:
						href = False
						for x in ["./h2/a","./div/h2/a"]:
							if not href:
								try: href = li.xpath(x)[0].get("href",False)
								except: pass
						urls.append(href)
				PAGE += 50
			except Exception as e:
				print e 
				isNot404 = False 
		for u in urls:
			r = self.http(u)
			art = {} 
			art['url'] = u 
			root = etree.fromstring(r,parser=self.parser)
			for x in ["/html/body/div[4]/header[1]/h1","/html/body/div[2]/header[1]/h1"]: 
				if 'name' not in art: 	
					try: 	art['name'] = root.xpath(x)[0].text
					except: pass
			for x in ['/html/body/div[2]/header[1]/dl/dd[2]',"/html/body/div[4]/header[1]/dl/dd[2]"]:
				if 'sold' not in art: 	
					try: 	art['sold'] = root.xpath(x)[0].text
					except: pass
			for x in ['//*[@id="productInfo"]/fieldset[1]/article/strong','//*[@id="productInfo"]/fieldset[1]/article/strong']:
				if 'value' not in art: 	
					try: 	art['value'] = root.xpath(x)[0].text
					except: pass
			for x in ['//*[@id="productInfo"]/fieldset[3]/article/p[2]/span','//*[@id="sellerInfo"]/header/dl/dd/span']:
				if 'place' not in art: 	
					try: 	art['place'] = root.xpath(x)[0].text
					except: pass
			for x in ['//*[@id="gallery_dflt"]/div[1]/a/img','//*[@id="gallery_dflt"]/div[1]/img',
				'//*[@id="productGalleryCollection"]/figure[1]/div[1]/img[1]','//*[@id="productGalleryCollection"]/figure[1]/div[1]/a/img[1]']:
				if 'img' not in art: 	
					try: 	art['img'] = root.xpath(x)[0].get('src') 
					except: pass
			art['value'] 	= int(art['value'].replace("$"," ").replace(".","").replace("\t",""))
			if 'sold' in art: 
				art['sold'] = int(art['sold'].replace("\t"," ").replace("vendidos","").replace(" ","").replace("vendido",""))
			else:
				art['sold'] = 0
			art['revenue'] 	= art['value'] * art['sold'] 
			results.append(art)	
		return results

	def orderBy(self,arr,col):
		i = 0 
		while i < len(arr):
			j = i+1
			while j < len(arr):
				if arr[i][col] < arr[j][col]: 
					arr[i],arr[j] = arr[j],arr[i]				
				j += 1
			i += 1 
		return arr	

def usage():
	print """	
	%s
	-s <keyword>	Search for this element at ML.
	-l <int>	Amount of elements to search for.
	-r <int> 	Amount of results to show.
	-p <path>	Path where output .xls will be created.
	-S		Show sold ranking.
	-R		Show revenue ranking.
	-V 		Show price value ranking.
	-o 		Open output file in gnome at the end.
	-h 		Display this help.
	""" % sys.argv[0]
	sys.exit(0)

if __name__ == "__main__": 
	try:
        	opts, args = getopt.getopt(sys.argv[1:],'s:p:r:l:SRVho', [])
        	myopt = {}
        	for opt, arg in opts:
               		myopt[opt] = arg or True
	except Exception as e:
        	print e
		usage()
	if "-h" in myopt: usage()	
	LIMIT 	= int(myopt.get("-l",150))
	RANKING = int(myopt.get("-r",10))
	OUTPATH = myopt.get("-p","/tmp")
	SEARCH_COLS = [] 
	if "-S" in myopt: SEARCH_COLS.append([1,"sold"])
	if "-V" in myopt: SEARCH_COLS.append([2,"value"])
	if "-R" in myopt: SEARCH_COLS.append([3,"revenue"])
	if "-s" in myopt: DICTIONARY = [myopt['-s']]
	else: DICTIONARY = [ 
			"Diseno Moderno",
			"Mesa diseno moderno",
			"mesa lcd",	
			"escritorio",
			"velador moderno",
			"estante moderno",
			"comoda",	
			"mueble computadora",
			"espejo",
			"modular",
			"modular moderno",
			"lampara diseno moderno",
			"lampara",
			"remera sublimadas", 
			"remera estampadas", 
	] 
	commands.getstatusoutput("mkdir -p %s" % OUTPATH)
	FILENAME = "%s/rho-ML-%i.xls" % (OUTPATH, time.time())
	ml = MercadoLibre()
	xls = openpyxl.Workbook()
	ws  = xls.worksheets[0]
	ws.cell(row=1,column=1)
	ws.cell(row=1,column=2)
	ws.cell(row=1,column=3)
	ws.cell(row=1,column=4)
	ws.cell(row=1,column=5)
	ws.cell(row=1,column=6)
	ws.column_dimensions["A"].width = 10
	ws.column_dimensions["B"].width = 10
	ws.column_dimensions["C"].width = 10
	ws.column_dimensions["D"].width = 55
	ws.column_dimensions["E"].width = 20
	ws.column_dimensions["F"].width = 100
	bold  = Style(font=Font(bold=True,size=13))
	right = Style(alignment=Alignment(horizontal='right'))
	row = 1 
	for s in DICTIONARY:
		r = ml.search(s,limit=LIMIT)
		for o in SEARCH_COLS:
			ws.cell(row=row,column=1).value = s.upper()
			ws.cell(row=row,column=1).style = bold
			row += 1 
			ws.cell(row=row,column=1).value = "VENDIDOS"
			ws.cell(row=row,column=2).value = "PRECIO"
			ws.cell(row=row,column=3).value = "VENTAS"
			ws.cell(row=row,column=4).value = "NOMBRE"
			ws.cell(row=row,column=5).value = "LUGAR"
			ws.cell(row=row,column=6).value = "URL"
			ws.cell(row=row,column=o[0]).style = bold
			row += 1 
			for e in ml.orderBy(r,o[1])[0:RANKING]: 
				ws.cell(row=row,column=1).value = "%i" 	% e['sold'] 
				ws.cell(row=row,column=2).value = "$%i" % e['value'] 
				ws.cell(row=row,column=3).value = "$%i" % e['revenue'] 
				ws.cell(row=row,column=4).value = e['name'] 
				ws.cell(row=row,column=5).value = e['place'] 
				ws.cell(row=row,column=6).value = e['url'] 
				ws.cell(row=row,column=2).style = right
				ws.cell(row=row,column=3).style = right
				row += 1 
			row += 1 
		xls.save(FILENAME)
	print FILENAME
	if "-o" in myopt:
		commands.getstatusoutput("gnome-open %s" % FILENAME)
