from openpyxl.styles import Style, Font, Alignment
from lxml import etree
import cookielib
import requests
import openpyxl
import commands 
import getopt
import rho
import time
import sys
import re

class Telexplorer(rho.RhoHTTP):

	def search(self,name,place,limit=40):
		PAGE 		= 1 
		results 	= []
		isNot404 	= True
		url 		= "http://buscador.telexplorer.com/ans.asp"
		loop		= 0	
		while len(results)<limit and isNot404 and loop<10:
			loop += 1
			try: 
				data = { 
					"q": "%s en %s" % (name,place),
					"pagina": PAGE, 
				} 
				r = self.http(url,data)
				root = etree.fromstring(r,parser=self.parser)
				for li in root.xpath('//*[@id="listado_resultados_residenciales"]/li'):
					loop = 0
					data = {}
					data['name'] 	= li.xpath('./p[1]/a')[0].text
					data['address'] = li.xpath('./p[3]/a')[0].text
					data['phone'] 	= li.xpath('./p[2]')[0].text
					data['city'] 	= li.xpath('./p[3]/text()')[0].encode('utf-8','ignore') 
					data['city'] 	= re.sub("^\s+","",data['city'])
					if len(results)<limit: 
						results.append(data)
				PAGE += 1
			except Exception as e:
				print e 
				isNot404 = False
		return results

def usage():
	print """	
	%s 
	-s <name>	Search name.
	-w <place>	Search location.
	-l <limit>	Amount of elements to retrieve.
	-h		Display help.
	-o 		Open .XLS file at the end.
	""" % sys.argv[0]
	sys.exit(0)

if __name__ == "__main__": 
	try:
        	opts, args = getopt.getopt(sys.argv[1:],'s:w:p:l:ho', [])
        	myopt = {}
        	for opt, arg in opts:
               		myopt[opt] = arg or True
	except Exception as e:
        	print e
		usage()
	if "-h" in myopt: usage()	
	LIMIT 	= int(myopt.get("-l",20))
	NAME	= myopt.get("-s")
	PLACE	= myopt.get("-w")
	OUTPATH = myopt.get("-p","/tmp")
	commands.getstatusoutput("mkdir -p %s" % OUTPATH)
	FILENAME = "%s/rho-TX-%i.xls" % (OUTPATH, time.time())
	tx = Telexplorer()
	xls = openpyxl.Workbook()
	ws  = xls.worksheets[0]
	ws.cell(row=1,column=1).value = "NAME"
	ws.cell(row=1,column=2).value = "PHONE"
	ws.cell(row=1,column=3).value = "ADDRESS"
	ws.cell(row=1,column=4).value = "CITY"
	ws.column_dimensions["A"].width = 40
	ws.column_dimensions["B"].width = 18
	ws.column_dimensions["C"].width = 30
	ws.column_dimensions["D"].width = 40
	bold  = Style(font=Font(bold=True,size=13))
	ws.cell(row=1,column=1).style = bold
	ws.cell(row=1,column=2).style = bold
	ws.cell(row=1,column=3).style = bold
	ws.cell(row=1,column=4).style = bold
	row = 2
	for e in tx.search(name=NAME,place=PLACE,limit=LIMIT):
		ws.cell(row=row,column=1).value = e['name'] 
		ws.cell(row=row,column=2).value = e['phone'] 
		ws.cell(row=row,column=3).value = e['address'] 
		ws.cell(row=row,column=4).value = e['city'] 
		row += 1 
	xls.save(FILENAME)
	print FILENAME
        if "-o" in myopt:
                commands.getstatusoutput("gnome-open %s" % FILENAME)
